import openpyxl
from typing import Dict, List, Tuple, Union
path = "precios.xlsx"

def precios() -> Dict[str, int]:
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    # print(ws['A1'].value)   # Nombre
    # print(ws.cell(row=1,column=1).value) # Nombre
    cant_filas = hoja.max_row
    productos = {}
    fila = 2
    while fila < cant_filas+1 and hoja.cell(row=fila, column=1).value:
        producto     = hoja.cell(row=fila, column=1).value
        precio       = hoja.cell(row=fila, column=2).value
        productos[producto] = precio
        fila += 1

    return productos

# print(precios()) # {'cafe': 30.97, 'oreo': 343.33, 'pan_lactal': 374.83}

def links() -> Dict[str, List[Tuple[str, int]]]:
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    productos = {}
    cant_filas = hoja.max_row
    fila = 2
    while fila < cant_filas+1 and hoja.cell(row=fila, column=1).value:
        producto     = hoja.cell(row=fila, column=1).value
        links        = hoja.cell(row=fila, column=3).value
        links = list(map(lambda x: tuple(x.split(";")), links.split(',')))
        productos[producto] = links
        fila += 1

    return productos

# print(links()) # {'cafe': [['https...dolca-suave-170-g-729427/p', '113'], ['https...la-virginia-bolsa-x-250-g-681613/p', '100']]}

def porcion() -> Dict[str, int]:
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    productos = {}
    cant_filas = hoja.max_row
    fila = 2
    while fila < cant_filas+1 and hoja.cell(row=fila, column=1).value:
        producto     = hoja.cell(row=fila, column=1).value
        porcion      = hoja.cell(row=fila, column=4).value
        productos[producto] = porcion
        fila += 1

    return productos

# print(porcion()) # {'cafe': 20, 'oreo': 60, 'pan_lactal': 200}

def unidad() -> Dict[str, str]:
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    productos = {}
    cant_filas = hoja.max_row
    fila = 2
    while fila < cant_filas+1 and hoja.cell(row=fila, column=1).value:
        producto     = hoja.cell(row=fila, column=1).value
        unidad       = hoja.cell(row=fila, column=5).value
        productos[producto] = unidad
        fila += 1

    return productos

# print(unidad()) # {'cafe': 'g, 'oreo': 'g', 'pan_lactal': 'g'}

def data() -> Dict[str, Dict[str, Union[float, str, List[Tuple[str, int]]]]]:
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    productos = {}
    cant_filas = hoja.max_row
    fila = 2
    while fila < cant_filas+1 and hoja.cell(row=fila, column=1).value:
        producto     = hoja.cell(row=fila, column=1).value
        precio       = hoja.cell(row=fila, column=2).value
        links        = hoja.cell(row=fila, column=3).value
        porcion      = hoja.cell(row=fila, column=4).value
        unidad       = hoja.cell(row=fila, column=5).value
        links = list(map(lambda x: x.split(";"), links.split(',')))
        productos[producto] = {'precio':precio, 'links':links, 'porcion':porcion, 'unidad':unidad}
        fila += 1

    return productos

# print(data()) # {'oreo': {'precio': 343.33, 'links': [('https://www.carrefour.com.ar/galletitas-dulce-oreo-rellenas-con-crema-118-g-715949/p', '3')], 'porcion': 60, 'unidad': 'g'}}

def data_producto(producto: str) -> Dict[str, Union[float, str, List[Tuple[str, int]]]]:
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    datos = {}
    cant_filas = hoja.max_row
    fila = 2
    while fila < cant_filas+1 and hoja.cell(row=fila, column=1).value:
        if producto == hoja.cell(row=fila, column=1).value:
            precio       = hoja.cell(row=fila, column=2).value
            links        = hoja.cell(row=fila, column=3).value
            porcion      = hoja.cell(row=fila, column=4).value
            unidad       = hoja.cell(row=fila, column=5).value
            links = list(map(lambda x: x.split(";"), links.split(',')))
            datos = {'precio':precio, 'links':links, 'porcion':porcion, 'unidad':unidad}
            fila += cant_filas
        fila += 1

    return datos

# print(data_producto('oreo')) # {'precio': 343.33, 'links': [('https://www.carrefour.com.ar/galletitas-dulce-oreo-rellenas-con-crema-118-g-715949/p', '3')], 'porcion': 60, 'unidad': 'g'}