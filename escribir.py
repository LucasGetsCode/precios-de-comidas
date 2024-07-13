import openpyxl
from typing import Dict, List, Tuple, Union
path = "precios.xlsx"

def modificar_producto(data: Dict[str, str]):
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    cant_filas = hoja.max_row
    fila = 2
    while fila < cant_filas+1 and hoja.cell(row=fila, column=1).value:
        if hoja.cell(row=fila, column=1).value == data['nombre']:
            hoja.cell(row=fila, column=1).value = data['nuevo_nombre']
            hoja.cell(row=fila, column=3).value = data['links']
            hoja.cell(row=fila, column=4).value = float(data['porcion'])
            hoja.cell(row=fila, column=5).value = data['unidad']
            fila += cant_filas
        fila += 1
    wb.save(path)

def agregar_producto(data: Dict[str,str]):
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    cant_filas = hoja.max_row
    fila = 2
    while fila < cant_filas+1 and hoja.cell(row=fila, column=1).value:
        if hoja.cell(row=fila, column=1).value == data['nombre']: return
        fila += 1
    hoja.cell(row=fila, column=1).value = data['nombre']
    hoja.cell(row=fila, column=3).value = data['links']
    hoja.cell(row=fila, column=4).value = float(data['porcion'])
    hoja.cell(row=fila, column=5).value = data['unidad']
    fila += cant_filas
    wb.save(path)