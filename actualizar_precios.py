from bs4 import BeautifulSoup
import openpyxl
import re
import concurrent.futures
import requests

path = "precios.xlsx"

def carrefour(pagina):
    if type(pagina) != BeautifulSoup:
        soup = BeautifulSoup(pagina.content, 'lxml')
    else:
        soup = pagina
    precio_hay_promo = soup.find('span', class_="valtech-carrefourar-product-price-0-x-listPrice")
    precio_no_hay_promo = soup.find('span', class_="valtech-carrefourar-product-price-0-x-sellingPriceValue")
    if precio_hay_promo: precio = precio_hay_promo.text
    else:                precio = precio_no_hay_promo.text
    precio = float(precio[2:].replace('.', '').replace(',', '.'))
    return precio


# print(carrefour(requests.get("https://www.carrefour.com.ar/cafe-molido-clasico-la-virginia-bolsa-x-250-g-681613/p")))
# print(carrefour(requests.get("https://www.carrefour.com.ar/nescafe-dolca-suave-170-g-729427/p")))
# print(coto(requests.get("https://www.cotodigital3.com.ar/sitios/cdigi/producto/-cafe-molido-torrado-equilibrado-la-virginia-paq-125-grm/_/A-00514019-00514019-200")))


def obtener_links():
    def links_celda(fila, columna):
        links_string = hoja.cell(row=fila, column=columna).value
        links_array = map(lambda x: x.split(';'), links_string.split(",")) if links_string else []
        return links_array
    
    wb = openpyxl.load_workbook(filename=path)
    hoja = wb.active

    # print(hoja['A1'].value)   # Nombre
    # print(hoja.cell(row=1,column=1).value) # Nombre
    productos = {}
    cant_filas = hoja.max_row
    for fila in range(2,cant_filas+1):
        producto           = hoja.cell(row=fila, column=1).value
        links_carrefour    = links_celda(fila, 3)
        celda_precio       = hoja.cell(row=fila, column=2)
        productos[producto] = {'celda_precio':celda_precio, 'fila':fila, 'links':links_carrefour, 'valores':0,'cant':0}
    return productos, wb

def actualizar_precios():
    def fetch_url(name, url_info):
        url, numero = url_info[0], url_info[1]
        try:
            response = requests.get(url)
            return name, url, numero, response.content
        except requests.RequestException as e:
            return name, url, numero, None

    data, archivo = obtener_links()  # { nombre : [[link,cant], [link,cant]], nombre : [[link,cant]...]}

    # Usando ThreadPoolExecutor para descargar en paralelo
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {}
        for name, info in data.items():
            for url_info in info['links']:
                futures[executor.submit(fetch_url, name, url_info)] = (name, url_info[0], url_info[1])
        
        for future in concurrent.futures.as_completed(futures):
            name, url, number = futures[future]
            try:
                name, url, number, html = future.result()
                if html:
                    print(f"Descargado {name} - {url} con éxito {number}")
                    resultados = carrefour(BeautifulSoup(html.decode('utf-8'), 'html.parser'))
                    data[name]['valores'] += (resultados / int(number))
                    data[name]['cant'] += 1
                else:
                    print(f"Error al descargar {name} - {url} {number}")
            except Exception as e:
                print(f"Excepción para {name} - {url}: {e}")
    for name, info in data.items():
        precio_porcion = info['valores']/info['cant']
        print(name + ' ' + str(precio_porcion))
        info['celda_precio'].value = round(precio_porcion,2)
    archivo.save(path)

# actualizar_precios()