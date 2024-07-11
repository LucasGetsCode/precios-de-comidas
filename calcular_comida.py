import obtener
import actualizar_precios
import openpyxl

path = "precios.xlsx"
wb = openpyxl.load_workbook(filename=path)
hoja = wb.active

def añadir_url(nombre, cant, url):
    ultima_fila = hoja.max_row
    productos = [hoja.cell(row=i,column=1).value for i in range(2,ultima_fila+1)]
    if nombre in productos:
        indice = productos.index(nombre)+2
        otros_links = hoja.cell(row=indice, column=3).value
        hoja.cell(row=indice, column=3, value=otros_links+','+url+';'+cant)
    else:
        hoja.cell(row=ultima_fila+1, column=1, value=nombre)
        hoja.cell(row=ultima_fila+1, column=3, value=url+';'+cant)
    print(productos)

def quitar_url(nombre, url):
    ultima_fila = hoja.max_row
    productos = [hoja.cell(row=i,column=1).value for i in range(2,ultima_fila+1)]
    if nombre in productos:
        indice = productos.index(nombre) + 2
        links = hoja.cell(row=indice, column=3).value.split(",")
        for i in reversed([j for j, elem in enumerate(links) if url == elem.split(';')[0]]):
            links.pop(i)
            print(f"Se quitó con éxito la {i}-ésima url")
    hoja.cell(row=indice, column=3, value=','.join(links))

def printear(nombre):
    ultima_fila = hoja.max_row
    productos = [hoja.cell(row=i,column=1).value for i in range(2,ultima_fila+1)]
    if nombre in productos:
        indice = productos.index(nombre) + 2
        precio = hoja.cell(row=indice, column=2).value
        final_url = lambda x: "http://..." + ';'.join([x.split(';')[0][-30:], x.split(';')[1]])
        links = map(final_url, hoja.cell(row=indice, column=3).value.split(','))
        print(f"Precio: {precio}. Links: {"  ".join(links)}")
    else:
        print("La palabra ingresada no es un producto existente ni se reconoce como palabra clave")


data = obtener.precios()

seguir = True
comidas = []
print("Sistema iniciado.")
while seguir:
    texto = input()
    partes = texto.split(" ")
    if texto == '' or texto == 'listo':
        seguir = False
    elif len(partes) == 3 and 'http' in partes[2]:
        añadir_url(partes[0], partes[1], partes[2])
    elif len(partes) == 2 and partes[1].replace('.','',1).isdigit():
        comidas.append((partes[0], partes[1]))
    elif len(partes) == 2 and 'http' in partes[1]:
        quitar_url(partes[0], partes[1])
    elif texto == "actualizar":
        print("Actualizando precios...")
        actualizar_precios.actualizar_precios()
        data = obtener.precios()
        print("Listo!")
    elif texto == "guardar":
        wb.save(path)
    elif texto == "calcular":
        total = 0
        for (comida, cant) in comidas:
            total += float(cant)*data[comida]['precio']
        print(f"El precio de las comidas ingresadas es ${round(total,2)}")
        comidas = []
    elif len(partes) == 1:
        printear(texto)